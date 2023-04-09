import openpyxl
from recipe_scrapers import scrape_me

# Scrape recipe from website
url = 'https://www.eatingwell.com/recipe/7910586/spinach-artichoke-dip-pasta-with-chicken/'
scraper = scrape_me(url, wild_mode=True)

# Get recipe information
image_url = scraper.image()
ingredients_list = scraper.ingredients()
try:
  total_time = scraper.total_time();
except:
  print("Total Time Not found :C")
  total_time = "N/A"
macros = scraper.nutrients()

# Create a new workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write headers
worksheet.cell(row=1, column=1, value='Image')
worksheet.cell(row=1, column=2, value='Ingredients')
worksheet.cell(row=1, column=3, value='Total Time')
worksheet.cell(row=1, column=4, value='Calories')
worksheet.cell(row=1, column=5, value='Fat')
worksheet.cell(row=1, column=6, value='Protein')
worksheet.cell(row=1, column=7, value='Carbohydrates')

# Write recipe information
worksheet.cell(row=2, column=1, value=image_url)
worksheet.cell(row=2, column=2, value='\n'.join(ingredients_list))
worksheet.cell(row=2, column=3, value= total_time)
worksheet.cell(row=2, column=4, value=macros['calories'])
worksheet.cell(row=2, column=5, value=macros['fat'])
worksheet.cell(row=2, column=6, value=macros['protein'])
worksheet.cell(row=2, column=7, value=macros['carbohydrates'])

# Save workbook
workbook.save('recipe_information.xlsx')
